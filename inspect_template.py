import openpyxl

filepath = r"C:\Users\崎久保秀一\Desktop\ClaudeWork\代理受領フォルダ\代理受領通知書_原本.xlsx"
outpath = r"C:\Users\崎久保秀一\Desktop\ClaudeWork\代理受領フォルダ\inspect_template_out.txt"

try:
    with open(outpath, 'w', encoding='utf-8') as f:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        if "原本" in wb.sheetnames:
            sheet = wb["原本"]
            f.write(f"Sheet name: {sheet.title}\n")
            for row in range(15, 23):
                val = sheet[f'H{row}'].value
                f.write(f"H{row}: {repr(val)}\n")
            
            # Check C25, D25
            f.write(f"C25: {repr(sheet['C25'].value)}\n")
            f.write(f"D25: {repr(sheet['D25'].value)}\n")
            f.write(f"E25: {repr(sheet['E25'].value)}\n")
        else:
            f.write("原本シートが見つかりません。\n")
            
        if "事業者情報" in wb.sheetnames:
            sheet = wb["事業者情報"]
            f.write(f"\nSheet name: {sheet.title}\n")
            for row in range(3, 9):
                val = sheet[f'C{row}'].value
                f.write(f"C{row}: {repr(val)}\n")
                
except Exception as e:
    print(f"Error reading excel: {e}")
