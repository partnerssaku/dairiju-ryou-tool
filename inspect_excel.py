import openpyxl
import os

folder = r"C:\Users\崎久保秀一\Desktop\ClaudeWork\代理受領フォルダ"
filename = "代理受領通知書_原本.xlsx"
filepath = os.path.join(folder, filename)
out_file = os.path.join(folder, "inspect_excel_out.txt")

if os.path.exists(filepath):
    with open(out_file, 'w', encoding='utf-8') as f:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            f.write(f"\n--- Sheet name: {sheet.title} ---\n")
            for row in sheet.iter_rows(max_row=60):
                for c in row:
                    if c.value is not None and str(c.value).strip() != "":
                        f.write(f'{c.coordinate}: {c.value}\n')
else:
    print(f"File not found: {filepath}")
