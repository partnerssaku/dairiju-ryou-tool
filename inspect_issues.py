import openpyxl

OUTPUT_FILE = r"C:\Users\崎久保秀一\Desktop\ClaudeWork\代理受領フォルダ\代理受領通知書_一括出力_20260224_190102.xlsx"
TEMPLATE_FILE = r"C:\Users\崎久保秀一\Desktop\ClaudeWork\代理受領フォルダ\代理受領通知書_原本.xlsx"

wb = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)

print("="*80)
print("ISSUE SCAN: Checking all 37 output sheets for problems")
print("="*80)

# Check key fields across ALL sheets
print("\n--- Per-sheet key cell values ---")
print(f"{'Sheet':<22s} {'D7(ID)':<14s} {'C25(mo)':<8s} {'E26(muni)':<12s} {'E27(svc)':<20s} {'E29(rcpt)':<16s} {'H30(svcA)':<10s} {'H31(burB)':<10s} {'H32(spcC)':<10s} {'H33(合計)':<10s} {'H34':<10s} {'F25(proxy)':<10s}")
print("-"*160)

issues = []
for sname in wb.sheetnames:
    ws = wb[sname]
    d7 = ws['D7'].value or ''
    c25 = ws['C25'].value
    e26 = ws['E26'].value or ''
    e27 = ws['E27'].value or ''
    e29 = ws['E29'].value or ''
    h30 = ws['H30'].value
    h31 = ws['H31'].value
    h32 = ws['H32'].value
    h33 = ws['H33'].value
    h34 = ws['H34'].value
    f25 = ws['F25'].value
    
    # row 24 vs row 25 check
    c24 = ws['C24'].value
    d24 = ws['D24'].value
    
    print(f"{sname:<22s} {str(d7):<14s} {str(c25):<8s} {str(e26):<12s} {str(e27):<20s} {str(e29):<16s} {str(h30):<10s} {str(h31):<10s} {str(h32):<10s} {str(h33):<10s} {str(h34):<10s} {str(f25):<10s}")
    
    # Check for issues
    if not e26 or e26.strip() == '':
        issues.append(f"  [{sname}] E26 (municipality) is EMPTY")
    if h31 == h30:
        issues.append(f"  [{sname}] H31={h31} equals H30={h30} -- H31 should be user burden (B), not total service cost")
    if h33 == 10000 and h32 != 10000:
        issues.append(f"  [{sname}] H33={h33} looks like special benefit was placed in 合計 row instead; H32={h32}")
    # Check formula: total = A - B + C => H34 should = H30 - H31 + H32
    if h30 is not None and h31 is not None and h32 is not None and h34 is not None:
        expected_total = h30 - h31 + h32
        if h34 != expected_total:
            issues.append(f"  [{sname}] H34={h34} != H30-H31+H32 = {expected_total}")
    # Row 24 still has template data
    if c24 == 9 and d24 == '月分サービス費':
        pass  # row 24 is from template, that's expected if template had it

print("\n" + "="*80)
print(f"ISSUES FOUND: {len(issues)}")
print("="*80)
for i in issues:
    print(i)

# Also compare template row layout vs output row layout
print("\n\n--- Template '原本' sheet: row mapping reference ---")
twb = openpyxl.load_workbook(TEMPLATE_FILE, data_only=True)
tws = twb['原本']
print("Row 24 (template):")
for c in range(1, 10):
    v = tws.cell(24, c).value
    if v is not None:
        print(f"  {tws.cell(24,c).coordinate} = {repr(v)}")
print("Row 25 (template):")
for c in range(1, 10):
    v = tws.cell(25, c).value
    if v is not None:
        print(f"  {tws.cell(25,c).coordinate} = {repr(v)}")
print("Rows 26-34 (template):")
for r in range(26, 35):
    for c in range(1, 10):
        v = tws.cell(r, c).value
        if v is not None:
            print(f"  {tws.cell(r,c).coordinate} = {repr(v)}")

print("\n--- Output first sheet 'ｱﾏﾉﾏｻｼ': row mapping ---")
ows = wb['ｱﾏﾉﾏｻｼ']
print("Row 24:")
for c in range(1, 10):
    v = ows.cell(24, c).value
    if v is not None:
        print(f"  {ows.cell(24,c).coordinate} = {repr(v)}")
print("Row 25:")
for c in range(1, 10):
    v = ows.cell(25, c).value
    if v is not None:
        print(f"  {ows.cell(25,c).coordinate} = {repr(v)}")
print("Rows 26-34:")
for r in range(26, 35):
    for c in range(1, 10):
        v = ows.cell(r, c).value
        if v is not None:
            print(f"  {ows.cell(r,c).coordinate} = {repr(v)}")

wb.close()
twb.close()
print("\nDone.")
