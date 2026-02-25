import openpyxl
import sys

OUTPUT_FILE = r"C:\Users\崎久保秀一\Desktop\ClaudeWork\代理受領フォルダ\代理受領通知書_一括出力_20260224_190102.xlsx"
TEMPLATE_FILE = r"C:\Users\崎久保秀一\Desktop\ClaudeWork\代理受領フォルダ\代理受領通知書_原本.xlsx"

def dump_sheet(ws, label, max_row=50, max_col=10):
    print(f"\n{'='*70}")
    print(f"  Sheet: [{label}] => '{ws.title}'")
    print(f"  Dimensions: {ws.dimensions}")
    print(f"{'='*70}")
    
    non_empty = []
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                addr = cell.coordinate
                val = cell.value
                # Show repr for strings to reveal line breaks
                if isinstance(val, str):
                    display = repr(val)
                else:
                    display = str(val)
                merged = ""
                # Check if cell is merged
                for mr in ws.merged_cells.ranges:
                    if cell.coordinate in mr:
                        merged = f"  [MERGED: {mr}]"
                        break
                non_empty.append((addr, display, merged))
    
    if non_empty:
        for addr, val, merged in non_empty:
            print(f"  {addr:6s} = {val}{merged}")
    else:
        print("  (no non-empty cells found in range)")
    
    # Also list merged cell ranges
    if ws.merged_cells.ranges:
        print(f"\n  Merged cell ranges ({len(ws.merged_cells.ranges)}):")
        for mr in sorted(ws.merged_cells.ranges, key=str):
            print(f"    {mr}")

def inspect_file(filepath, label):
    print(f"\n{'#'*70}")
    print(f"# FILE: {label}")
    print(f"# Path: {filepath}")
    print(f"{'#'*70}")
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    print(f"\nSheet names: {wb.sheetnames}")
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        dump_sheet(ws, f"{label} / {sname}")
    
    wb.close()
    return wb

# ---- Inspect Template ----
inspect_file(TEMPLATE_FILE, "TEMPLATE")

# ---- Inspect Output ----
print("\n\n")
inspect_file(OUTPUT_FILE, "OUTPUT")

# ---- Compare first user sheet against expected mapping ----
print(f"\n\n{'#'*70}")
print("# COMPARISON: First user sheet vs expected mapping")
print(f"{'#'*70}")

wb = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)
sheets = wb.sheetnames
# The first user sheet (skip any sheet named like the template base)
# Typically the first sheet or first non-template sheet
user_sheet = wb[sheets[0]]
print(f"\nUsing sheet: '{user_sheet.title}'")

expected_mapping = {
    "D7":  ("user ID", None),
    "D8":  ("display name", None),
    "H4":  ("issue date", None),
    "C25": ("month number", None),
    "D25": ("月分サービス費", None),
    "E26": ("municipality", None),
    "E27": ("service type", None),
    "E29": ("receipt date", None),
    "F25": ("proxy amount", None),
    "H30": ("total service cost", None),
    "H31": ("total service cost (copy)", None),
    "H32": ("user burden", None),
    "H33": ("special benefit", None),
    "H34": ("total", None),
}

company_cells = ["H15", "H16", "H17", "H18", "H19"]

print("\n--- Expected mapping cells ---")
for addr, (desc, _) in expected_mapping.items():
    cell = user_sheet[addr]
    val = cell.value
    if isinstance(val, str):
        display = repr(val)
    else:
        display = str(val)
    status = "EMPTY!" if val is None else "OK"
    print(f"  {addr:6s} ({desc:30s}) = {display:50s}  [{status}]")

print("\n--- Company info cells (H15-H19, check line breaks) ---")
for addr in company_cells:
    cell = user_sheet[addr]
    val = cell.value
    if isinstance(val, str):
        display = repr(val)
        has_newline = "\n" in val
        print(f"  {addr:6s} = {display}")
        print(f"         has line breaks: {has_newline}")
    elif val is None:
        print(f"  {addr:6s} = None  [EMPTY]")
    else:
        print(f"  {addr:6s} = {val}")

# Check all sheets in output
print("\n--- Summary of all output sheets ---")
for sname in sheets:
    ws = wb[sname]
    # Count non-empty cells in key area
    count = 0
    for row in range(1, 40):
        for col in range(1, 10):
            if ws.cell(row=row, column=col).value is not None:
                count += 1
    print(f"  Sheet '{sname}': {count} non-empty cells in A1:I39")

wb.close()
print("\nDone.")
