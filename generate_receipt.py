import csv
import os
import glob
from openpyxl import load_workbook
import datetime

def find_latest_csv(folder_path):
    files = glob.glob(os.path.join(folder_path, '*.CSV'))
    if not files:
        return None
    latest_file = max(files, key=os.path.getctime)
    return latest_file

def parse_csv(csv_path):
    users = {}
    with open(csv_path, 'r', encoding='shift_jis', errors='replace') as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) < 3:
                continue
            record_id = row[2]
            
            # UKE J131 (01) レコードから金額情報を抽出
            if record_id == 'J131' and len(row) > 3 and row[3] == '01':
                user_id = row[7]
                kana_name = row[9]
                target_month = row[4] # YYYYMM
                
                try:
                    cost = int(row[22]) if len(row) > 22 else 0
                    burden = int(row[23]) if len(row) > 23 else 0
                    proxy = int(row[29]) if len(row) > 29 else 0
                    spec_grant = int(row[35]) if len(row) > 35 else 0
                except ValueError:
                    cost, burden, proxy, spec_grant = 0, 0, 0, 0
                
                key = (user_id, target_month)
                if key not in users:
                    users[key] = {
                        'user_id': user_id,
                        'name': kana_name,
                        'month': target_month,
                        'service_cost': cost,
                        'user_burden': burden,
                        'proxy_amount': proxy,
                        'spec_grant': spec_grant,
                        'service_code': ''
                    }
                else:
                    users[key]['service_cost'] += cost
                    users[key]['user_burden'] += burden
                    users[key]['proxy_amount'] += proxy
                    users[key]['spec_grant'] += spec_grant
            
            # UKE J131 (03) レコードからサービス種類コードを抽出
            elif record_id == 'J131' and len(row) > 3 and row[3] == '03':
                user_id = row[7]
                target_month = row[4]
                service_code = row[8] if len(row) > 8 else ''
                
                key = (user_id, target_month)
                if key in users and users[key]['service_code'] == '':
                    # 最初のサービスコードだけ保持する（またはすべて結合するなど）
                    # 今回は代表的な1つ目を取得
                    users[key]['service_code'] = service_code
                    
    # Flatten the dict if a user only has one month, otherwise append month to name
    results = {}
    for (user_id, month), data in users.items():
        base_name = data['name']
        # If there's multiple months for same user, we'd need to distinguish them
        key_str = f"{user_id}_{month}"
        results[key_str] = data
        
    return list(results.values())

def extract_masters(wb):
    masters = {
        'company': {},
        'municipality': {},
        'users': {}
    }
    
    # 1. 事業者情報の抽出
    if "事業者情報" in wb.sheetnames:
        ws = wb["事業者情報"]
        masters['company'] = {
            'name': ws['C3'].value if ws['C3'].value else "",
            'office': ws['C4'].value if ws['C4'].value else "",
            'rep_title': ws['C5'].value if ws['C5'].value else "",
            'rep_name': ws['C6'].value if ws['C6'].value else "",
            'tel': ws['C7'].value if ws['C7'].value else "",
            'fax': ws['C8'].value if ws['C8'].value else ""
        }
        
    # 2. 受給者情報の抽出 (受給者番号をキーにして、市町村・漢字氏名を保持)
    if "受給者情報" in wb.sheetnames:
        ws = wb["受給者情報"]
        # 左側の表
        for row_idx in range(3, 30):
            muni = ws.cell(row=row_idx, column=1).value
            uid = str(ws.cell(row=row_idx, column=2).value).strip() if ws.cell(row=row_idx, column=2).value else None
            name = ws.cell(row=row_idx, column=3).value
            if uid and uid != 'None':
                masters['users'][uid] = {'municipality': muni, 'kanji_name': name}
        
        # 右側の表
        for row_idx in range(3, 30):
            muni = ws.cell(row=row_idx, column=5).value
            uid = str(ws.cell(row=row_idx, column=6).value).strip() if ws.cell(row=row_idx, column=6).value else None
            name = ws.cell(row=row_idx, column=7).value
            if uid and uid != 'None':
                masters['users'][uid] = {'municipality': muni, 'kanji_name': name}
                
    return masters

def main():
    folder_path = r"C:\Users\崎久保秀一\Desktop\ClaudeWork\代理受領フォルダ"
    template_path = os.path.join(folder_path, "代理受領通知書_原本.xlsx")
    
    if not os.path.exists(template_path):
        print(f"エラー: テンプレートファイルが見つかりません: {template_path}")
        return

    csv_path = find_latest_csv(folder_path)
    if not csv_path:
        print(f"エラー: フォルダ内にCSVファイルが見つかりません。")
        return
        
    print(f"読み込みCSV: {os.path.basename(csv_path)}")
    
    today_str = datetime.datetime.now().strftime("%Y年%m月%d日")
    issue_date = input(f"発行日を入力してください（そのままEnterで「{today_str}」になります）: ")
    if not issue_date.strip():
        issue_date = today_str
        
    receipt_date = input(f"受給日を入力してください（例: 令和7年11月20日）: ")
    
    # ユーザーが指定するサービス名
    service_name_input = input(f"サービス種別を入力してください（例: 就労継続支援Ｂ型 など）: ")

    print("データを解析しています...")
    users_data = parse_csv(csv_path)
    
    if not users_data:
        print("エラー: CSVから対象データを抽出できませんでした。")
        return
        
    print(f"対象者: {len(users_data)} 件のデータが見つかりました。")
    print("Excelファイルを作成しています...")

    # Excelテンプレートの読み込み (数式ではなく値として処理するためマスターをPythonで展開)
    wb = load_workbook(template_path)
    wb_data = load_workbook(template_path, data_only=True)
    
    masters = extract_masters(wb)
    user_masters = masters['users']
    
    if "原本" in wb.sheetnames:
        source_sheet = wb["原本"]
        source_data = wb_data["原本"]
    else:
        source_sheet = wb.worksheets[0]
        source_data = wb_data.worksheets[0]
        
    company_vals = {
        'H15': source_data['H15'].value,
        'H16': source_data['H16'].value,
        'H17': source_data['H17'].value,
        'H18': source_data['H18'].value,
        'H19': source_data['H19'].value,
    }
        
    sheet_count = 0
    for i, data in enumerate(users_data):
        user_id = str(data['user_id']).strip()
        
        # マスターから漢字氏名と市町村を取得
        kanji_name = ""
        municipality = ""
        if user_id in user_masters:
            kanji_name = user_masters[user_id]['kanji_name']
            municipality = user_masters[user_id]['municipality']
            
        # 漢字氏名があれば優先、なければカナ
        display_name = kanji_name if kanji_name else data['name']
        
        safe_name = display_name[:31].replace(':', '').replace('\\', '').replace('/', '').replace('?', '').replace('*', '').replace('[', '').replace(']', '')
        if not safe_name:
            safe_name = f"User_{i}"
            
        ws = wb.copy_worksheet(source_sheet)
        ws.title = safe_name
            
        # 基本情報
        ws['D7'] = user_id
        ws['D8'] = display_name
        ws['H4'] = f"発行日: {issue_date}"
        
        # 受給日の書き込み (E29セル) -> 新レイアウトではずれているかもしれないが元の指定通りE29で維持
        ws['E29'] = receipt_date
        
        # マスター情報（Python側で直接文字を書き込むので、#REF!エラーにならない）
        ws['E26'] = municipality  # 市町村
        
        # サービス種類コードの変換（よくある種類を定義）
        SERVICE_NAMES = {
            '33': '就労継続支援Ｂ型',
            '32': '就労継続支援Ａ型',
            '43': '共同生活援助',
        }
        
        # ユーザー様が指定した場合はその文字優先、未入力の場合はコードから推測
        if service_name_input.strip():
            service_name = service_name_input
        else:
            raw_code = data.get('service_code', '')
            prefix = raw_code[:2] if raw_code else ''
            service_name = SERVICE_NAMES.get(prefix, raw_code)
            
        ws['E27'] = service_name  # サービス種類
        
        # 原本シートに用意していただいたレイアウト・文字をそのまま復元する
        ws['H15'] = company_vals.get('H15', '')
        ws['H16'] = company_vals.get('H16', '')
        ws['H17'] = company_vals.get('H17', '')
        ws['H18'] = company_vals.get('H18', '')
        ws['H19'] = company_vals.get('H19', '')

        # 対象月のフォーマット (YYYYMM -> YYYY年MM月)
        month_str = data['month']
        if len(month_str) == 6:
            year, month = month_str[:4], month_str[4:]
            ws['C25'] = int(month)  # 例: 1
            ws['D25'] = "月分サービス費"  # テンプレートによってはD25が空白化しているかもしれないので補填
        
        # 金額関係
        proxy_total = data['proxy_amount']
        ws['F25'] = proxy_total           # 代理受領額
        ws['H30'] = data['service_cost']  # 単位数
        ws['H31'] = data['service_cost']  # 総サービス費(A)
        ws['H32'] = data['user_burden']   # 利用者負担額(B)
        ws['H33'] = data['spec_grant']    # 特定障害者給付費(C)
        ws['H34'] = proxy_total           # 合計
        
        sheet_count += 1
        print(f" - {safe_name} 様 ({user_id}) のシートを作成中... ({sheet_count}/{len(users_data)})")

    # 全ての不要なシート（原本、マスターなど）を削除し、純粋な完成シートのみ残す
    # 削除することで1クリックの完全な状態になる
    valid_titles = []
    for ws in wb.worksheets:
        valid_titles.append(ws.title)
        
    for title in wb.sheetnames:
        if title in ["事業者情報", "受給者情報", "原本", "Sheet1"]:
            del wb[title]

    # 保存処理
    output_filename = f"代理受領通知書_一括出力_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(folder_path, output_filename)
    wb.save(output_path)
    
    print("\n" + "="*40)
    print("◆◆ 処理が完了しました！ ◆◆")
    print(f"出力ファイル: {output_filename}")
    print(f"作成シート数: {sheet_count}名分")
    print("="*40 + "\n")

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()
    input("Enterキーを押して画面を閉じます...")
